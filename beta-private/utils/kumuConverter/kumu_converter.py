from openpyxl import load_workbook
import sys, getopt

elements_sheet_name = 'Elements'
connections_sheet_name = 'Connections'
sheet_names = [elements_sheet_name, connections_sheet_name]

# These are the headings expected to be found in the KUMU Input file's 'Elements' sheet.
elements_sheet_headings = [
    'ID',
    'Label',
    'Version',
    'Type',
    'URL',
    'Description',
    'Tags',
    'Capability Type',
    'Capability Specific Standard 1',
    'Capability Specific Standard URL 1',
    'Capability Specific Standard 2',
    'Capability Specific Standard URL 2'
]

connections_sheet_headings = [
    'From',
    'To',
    'FromID',
    'ToID',
    'Type'
]

mappings_output_headings = [
    'CapabilityId',
    'StandardId',
    'IsOptional'
]

mappings_headings_map = {
    'CapabilityId': 'FromID',
    'StandardId': 'ToID',
    'IsOptional': 'Type',
}

# These are the output headings for the 'standards.csv' file
standard_output_headings = [
    'Id',
    'PreviousId',
    'IsOverarching',
    'Name',
    'Description',
    'URL',
    'Type',
]

# these are the mappings of output headings to kumu input headings for the 'standards.csv' file
standard_headings_map = {
    'Id': 'ID',
    'PreviousId': None,
    'IsOverarching': 'Type',
    'Name': 'Label',
    'Description': 'Description',
    'URL': 'URL',
    'Type': 'Type'
}

# These are the output headings for the 'capabilities.csv' file
capability_output_headings = [
    'Id',
    'PreviousId',
    'Name',
    'Description',
    'URL',
    'Type'
]

# These are the mappings of output headings to Kumu input headings for the 'capabilities.csv' file.
capability_headings_map = {
    'Id': 'ID',
    'PreviousId': None,
    'Name': 'Label',
    'Description': 'Description',
    'URL': 'URL',
    'Type': 'Capability Type',
}

def map_capability_type (row):
    '''Based on the capability output headings, convert the type from "Foundation" to "C" or "N".'''
    idx = capability_output_headings.index('Type')
    row[idx] = 'C' if row[idx] == 'Foundation' else 'N'
    return row


def map_standard_type (row):
    '''Based on the standard output headings, convert the kumu type into the catalogue types'''
    idx = standard_output_headings.index('Type')

    if 'Context Specific' in row[idx]:
        row[idx] = 'X'
    elif 'Capability Specific' in row[idx]:
        row[idx] = 'C'
    elif 'Overarching' in row[idx]:
        row[idx] = 'O'
    elif 'Interop' in row[idx]:
        row[idx] = 'I'

    return row


def map_standard_is_overarching (row):
    '''Based on the standard type, set the isOverarching flag.'''
    idx = standard_output_headings.index('IsOverarching')
    row[idx] = 1 if 'Overarching' in row[idx] else 0
    return row


def map_mapping_is_optional (row):
    '''none of these are optional, so just map to 0'''
    idx = mappings_output_headings.index('IsOptional')
    if 'Inherited' in row[idx]:
        row[idx] = 1
    else:
        row[idx] = 0
    return row


def load_work_book (fp):
    '''loads workbook into memory'''
    return load_workbook(fp)


def get_sheet_names (wb):
    '''gets an array of the sheetnames found in the workbook'''
    return wb.sheetnames


def get_sheet_headings (ws):
    '''gets headings from the specified sheet in the provided work_book'''
    iter_rows = list(ws.iter_rows())[0]
    return [cell.value for cell in iter_rows]


def valid_sheet_names (wb):
    '''checks if the sheetnames of the provided workbook match the expected sheet_names'''
    return get_sheet_names(wb) == sheet_names


def valid_elements_sheet_headings (ws):
    '''checks that the headings of the "Elements" sheet are as expected'''
    return list(filter(lambda x: x is not None, get_sheet_headings(ws))) == elements_sheet_headings


def valid_connections_sheet_headings (ws):
    '''checks that the headings of the "Elements" sheet are as expected'''
    return list(filter(lambda x: x is not None, get_sheet_headings(ws))) == connections_sheet_headings


def validate_work_book (wb):
    '''validates that the workbook is in an expected format'''
    elements_ws = wb.get_sheet_by_name(elements_sheet_name)
    connections_ws = wb.get_sheet_by_name(connections_sheet_name)

    if not valid_sheet_names(wb):
        print('Invalid Sheet names. Expected:', sheet_names, 'Recieved:', get_sheet_names(wb))
        return False
    if not valid_elements_sheet_headings(elements_ws):
        print('Invalid Sheet Headings for', elements_sheet_name, 'Sheet.\nExpected:', elements_sheet_headings, '\nRecieved:', get_sheet_headings(elements_ws))
        return False
    if not valid_connections_sheet_headings(connections_ws):
        print('Invalid Sheet Headings for', connections_sheet_name, 'Sheet.\nExpected:', connections_sheet_headings, '\nRecieved:', get_sheet_headings(connections_ws))
        return False
    return True


def get_col_of_mapped_headings (ws, heading_map):
    headings = get_sheet_headings(ws)
    return [
        headings.index(heading_map[output_heading]) if heading_map[output_heading] is not None else None for output_heading in heading_map
    ]


def get_col_of_heading (ws, heading):
    headings = get_sheet_headings(ws)
    return headings.index(heading)


def extract_capability_rows (ws):
    filter_heading_index = get_col_of_heading(ws, 'Type')
    capability_rows = []
    for row in ws:
        if list(row)[filter_heading_index].value == 'Capability':
            capability_rows.append(row)
    return capability_rows


def build_capabilities_output (wb):
    ws = wb.get_sheet_by_name(elements_sheet_name)

    capability_rows = extract_capability_rows(ws)
    heading_indices = get_col_of_mapped_headings(ws, capability_headings_map)

    csv_rows = [
        [list(row)[idx].value if idx is not None else '' for idx in heading_indices] for row in capability_rows
    ]

    csv_rows = [map_capability_type(row) for row in csv_rows]
    csv_rows.insert(0, capability_output_headings)

    return csv_rows


def extract_standard_rows (ws):
    filter_heading_index = get_col_of_heading(ws, 'Type')
    standard_rows = []
    for row in ws:
        row_val = list(row)[filter_heading_index].value
        not_none = row_val != None
        if not_none and 'Standard' in row_val:
            standard_rows.append(row)
    return standard_rows


def build_standards_output (wb):
    ws = wb.get_sheet_by_name(elements_sheet_name)

    standard_rows = extract_standard_rows(ws)
    heading_indices = get_col_of_mapped_headings(ws, standard_headings_map)

    csv_rows = [
        [list(row)[idx].value if idx is not None else '' for idx in heading_indices] for row in standard_rows
    ]

    csv_rows = [map_standard_type(row) for row in csv_rows]
    csv_rows = [map_standard_is_overarching(row) for row in csv_rows]
    csv_rows.insert(0, standard_output_headings)

    return csv_rows


def extract_capability_standard_rows (ws):
    from_col_idx = get_col_of_heading(ws, 'FromID')
    to_col_idx = get_col_of_heading(ws, 'ToID')
    type_col_idx = get_col_of_heading(ws, 'Type')

    capability_standard_rows = []
    for row in ws:
        from_val =  list(row)[from_col_idx].value
        to_val = list(row)[to_col_idx].value
        type_val = list(row)[type_col_idx].value

        not_none = (from_val != None) and (to_val != None)
        if not_none and from_val.startswith('C') and to_val.startswith('S'):
            capability_standard_rows.append(row)
    return capability_standard_rows


def build_mappings_output (wb):
    ws = wb.get_sheet_by_name(connections_sheet_name)

    capability_standards_rows = extract_capability_standard_rows(ws)
    heading_indices = get_col_of_mapped_headings(ws, mappings_headings_map)

    csv_rows = [
        [list(row)[idx].value if idx is not None else '' for idx in heading_indices] for row in capability_standards_rows
    ]

    csv_rows = [map_mapping_is_optional(row) for row in csv_rows]
    csv_rows.insert(0, mappings_output_headings)
    return csv_rows


def output_to_csv(contents, fp, seperator = '\t'):
    f = open(fp, 'w')
    for row in contents:
        row = [str(val) for val in row]
        f.write(seperator.join(row)+'\n')


def print_help ():
    '''Prints the command line options'''


def main (argv):

    in_fp = '../Kumu-gp-it-futures-capability-map v5.xlsx'
    out_cap_fp = 'Capabilities.tsv'
    out_std_fp = 'Standards.tsv'
    mapping_std_fp = 'CapabilityStandard.tsv'
    delimiter = '\t'

    try:
        opts, args = getopt.getopt(argv, 'hi:csd', ['input=', 'capability-output=', 'standard-output=', 'delimiter='])
    except getopt.GetoptError:
        print('kumu_converter.py -i <input-file-path> -c <capability-output-file> -s <standards-output-file> -m <mapping-output-file> -d <delimiter>')
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print('kumu_converter.py -i <input-file-path> -c <capability-output-file> -s <standards-output-file> -m <mapping-output-file> -d <delimiter>')
            print('-i, --input\t\t is required and specifies what file is being input to the program.')
            print('-c, --capability-output\t can be used to name the file that capabilities are output to.')
            print('-s, --standard-output\t can be used to name the file that standards are output to.')
            print('-m, --mapping-output\t can be used to name the file that capability-standard mappings are output to.')
            print('-d, --delimiter\t\t can be used to customise the delimiter used in output files.')
            sys.exit()
        elif opt in ('-i', '--input'):
            in_fp = arg
        elif opt in ('-c', '--capability-output'):
            out_cap_fp = arg
        elif opt in ('-s', '--standards-output'):
            out_std_fp = arg
        elif opt in ('-m', '--mapping-output'):
            mapping_std_fp = arg
        elif opt in ('-d', '--delimiter'):
            delimiter = arg

    wb = load_work_book(in_fp)
    valid_workbook = validate_work_book(wb)

    if not valid_workbook:
        print('Invalid workbook provided. Exiting.')
        return


    capability_output = build_capabilities_output(wb)
    standards_output = build_standards_output(wb)
    mappings_output = build_mappings_output(wb)

    output_to_csv(capability_output, out_cap_fp, delimiter)
    output_to_csv(standards_output, out_std_fp, delimiter)
    output_to_csv(mappings_output, mapping_std_fp, delimiter)

if __name__ == '__main__':
    main(sys.argv[1:])