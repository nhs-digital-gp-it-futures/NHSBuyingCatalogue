from openpyxl import load_workbook

elements_sheet_name = 'Elements'
connections_sheet_name = 'Connections'
sheet_names = [elements_sheet_name, connections_sheet_name]

elements_sheet_headings = [
    'ID',
    'Label',
    'Version',
    'Type',
    'URL',
    'Description',
    'Capability Type',
    'Tags',
    'Capability Specific Standard 1',
    'Capability Specific Standard URL 1',
    'Capability Specific Standard 2',
    'Capability Specific Standard URL 2'
]


def load_work_book (fp):
    '''loads workbook into memory'''
    return load_workbook(fp)

def get_sheet_names (wb):
    '''gets an array of the sheetnames found in the workbook'''
    return wb.sheetnames

def get_sheet_headings (wb, sheet_name):
    '''gets headings from the specified sheet in the provided work_book'''
    iter_rows = list(wb.get_sheet_by_name(elements_sheet_name).iter_rows())[0]
    return [cell.value for cell in iter_rows]

def valid_sheet_names (wb):
    '''checks if the sheetnames of the provided workbook match the expected sheet_names'''
    return get_sheet_names(wb) == sheet_names

def valid_elements_sheet_headings (wb):
    '''checks that the headings of the "Elements" sheet are as expected'''
    return get_sheet_headings(wb, elements_sheet_name) == elements_sheet_headings

def validate_work_book (wb):
    '''validates that the workbook is in an expected format'''
    if not valid_sheet_names(wb):
        print('Invalid Sheet names. Expected:', sheet_names, 'Recieved:', get_sheet_names(wb))
        return False
    if not valid_elements_sheet_headings(wb):
        print('Invalid Sheet Headings for', elements_sheet_name, 'Sheet.\nExpected:', elements_sheet_headings, '\nRecieved:', get_sheet_headings(wb, elements_sheet_name))
        return False
    return True

def extract_standards ()

def main ():
    wb = load_work_book('../Kumu-gp-it-futures-capability-map v5.xlsx')
    valid_workbook = validate_work_book(wb)

    if not valid_workbook:
        print('Invalid workbook provided. Exiting.')
        return



if __name__ == '__main__':
    main()